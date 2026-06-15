"""
Render the infra cost report as a branded PDF (AI Identity Purple) + an .xlsx
workbook with month-over-month tracking. Imported by infra_cost_report.py.
"""

import re

from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

# ── AI Identity Purple theme (matches docs/build_ocsf_brief.py) ──────────────
PURPLE_DEEP = HexColor("#4B2D7F")
PURPLE_MID = HexColor("#6B4FA0")
LAVENDER = HexColor("#F0E8F8")
WHITE = HexColor("#FFFFFF")
INK = HexColor("#1A1F36")
GREY = HexColor("#6B7280")
LINE = HexColor("#E2E0EE")
GREEN = HexColor("#1F9D55")
AMBER = HexColor("#C77700")

REG, BOLD, ITAL = "Helvetica", "Helvetica-Bold", "Helvetica-Oblique"
PAGE_W, PAGE_H = letter  # 612 x 792 (portrait)
LEFT = 50
CW = PAGE_W - 2 * LEFT


def _rect(c, x, top, w, h, fill=None, stroke=None, sw=0.6, radius=None):
    if fill:
        c.setFillColor(fill)
    if stroke:
        c.setStrokeColor(stroke)
        c.setLineWidth(sw)
    y = PAGE_H - top - h
    if radius:
        c.roundRect(x, y, w, h, radius, stroke=1 if stroke else 0, fill=1 if fill else 0)
    else:
        c.rect(x, y, w, h, stroke=1 if stroke else 0, fill=1 if fill else 0)


def _text(c, x, top, s, font=REG, size=10, color=INK, center_w=None):
    c.setFont(font, size)
    c.setFillColor(color)
    if center_w is not None:
        x = x + (center_w - stringWidth(s, font, size)) / 2
    c.drawString(x, PAGE_H - top - size, s)


def _wrap(text, font, size, max_w):
    out = []
    for raw in str(text).split("\n"):
        cur = ""
        for w in raw.split(" "):
            t = w if not cur else cur + " " + w
            if stringWidth(t, font, size) <= max_w:
                cur = t
            else:
                if cur:
                    out.append(cur)
                cur = w
        out.append(cur)
    return out


def build_pdf(label, generated, summary, flags, neon, atlas, gcp, sentry, path):
    c = canvas.Canvas(str(path), pagesize=letter)

    # ── header band ──
    _rect(c, 0, 0, PAGE_W, 92, fill=PURPLE_DEEP)
    _rect(c, 0, 92, PAGE_W, 4, fill=PURPLE_MID)
    _text(c, LEFT, 20, "AI IDENTITY", BOLD, 11, LAVENDER)
    _text(c, LEFT, 36, "Infrastructure Cost Report", BOLD, 22, WHITE)
    _text(c, LEFT, 66, f"{label}   ·   generated {generated}", REG, 10, LAVENDER)

    y = 120

    # ── executive summary band ──
    _rect(c, LEFT, y, CW, 58, fill=LAVENDER, radius=6)
    _text(c, LEFT + 16, y + 12, "EXECUTIVE SUMMARY", BOLD, 9, PURPLE_DEEP)
    big = f"${summary.get('out_of_pocket_usd', 0):,.2f}"
    _text(c, LEFT + 16, y + 26, big, BOLD, 20, INK)
    _text(c, LEFT + 16, y + 50, "cash out-of-pocket this month — rest on credits", REG, 9, GREY)
    # right side: credit runway
    rx = LEFT + CW - 230
    _text(c, rx, y + 12, "GCP CREDIT RUNWAY", BOLD, 9, PURPLE_DEEP)
    _text(c, rx, y + 26, f"${summary.get('gcp_credits_total', 0):,.0f}", BOLD, 20, INK)
    _text(
        c,
        rx,
        y + 50,
        f"{summary.get('gcp_credit_pools', 0)} pools · Atlas M10 on credits",
        REG,
        9,
        GREY,
    )
    y += 78

    # ── action items / flags ──
    _text(c, LEFT, y, "Action Items & Flags", BOLD, 13, PURPLE_DEEP)
    y += 22
    for f in flags:
        dot = GREEN if "✅" in f else AMBER
        body = re.sub(r"^[^0-9A-Za-z]+", "", f).strip()  # strip leading emoji + spaces
        c.setFillColor(dot)
        c.circle(LEFT + 4, PAGE_H - y - 4, 3.2, fill=1, stroke=0)
        lines = _wrap(body, REG, 9.5, CW - 22)
        for i, ln in enumerate(lines):
            _text(c, LEFT + 16, y + i * 12 - 4, ln, REG, 9.5, INK)
        y += max(14, len(lines) * 12)
    y += 10

    # ── per-service cards ──
    def card(title, emoji, rows, top, h):
        _rect(c, LEFT, top, CW, h, fill=WHITE, stroke=LINE, sw=0.8, radius=6)
        _rect(c, LEFT, top, 4, h, fill=PURPLE_MID)
        _text(c, LEFT + 16, top + 12, title, BOLD, 12, PURPLE_DEEP)
        ry = top + 32
        for k, v in rows:
            _text(c, LEFT + 16, ry, k, REG, 9.5, GREY)
            _text(c, LEFT + 180, ry, v, BOLD, 9.5, INK)
            ry += 15
        return top + h + 12

    # Neon
    neon_rows = [
        ("Plan", str(neon.get("plan", "?")).title()),
        ("Projects", str(len(neon.get("projects", [])))),
        ("Storage", f"{neon.get('total_storage_mb', 0):.1f} MB"),
        ("Monthly cost", "$0.00 (Free tier)"),
    ]
    y = card("Neon — Postgres", "🐘", neon_rows, y, 32 + len(neon_rows) * 15 + 8)

    # Atlas
    atlas_rows = [
        ("Cluster tier", ", ".join(atlas.get("tiers", [])) + " (GCP)"),
        ("This month (MTD)", f"${atlas.get('pending_usd', 0):,.2f}  (billed $0 — credits)"),
        ("Run-rate (last full mo.)", f"${atlas.get('last_invoice_usd', 0):,.2f}"),
    ]
    y = card("MongoDB Atlas", "🍃", atlas_rows, y, 32 + len(atlas_rows) * 15 + 8)

    # GCP
    gcp_rows = [("Billing method", str(gcp.get("method", "?")).title())]
    for cr in gcp.get("credits", []):
        gcp_rows.append((cr["name"], f"${cr['usd']:,.2f}  (exp {cr['expiry']})"))
    y = card("GCP / GKE — Autopilot", "☸️", gcp_rows, y, 32 + len(gcp_rows) * 15 + 8)

    # Sentry
    if sentry.get("configured"):
        cash = {"team": 29.0, "business": 80.0}.get((sentry.get("plan") or "").lower(), 0.0)
        s_rows = [
            ("Plan", str(sentry.get("plan", "?")).title()),
            ("Projects", str(len(sentry.get("projects", [])))),
            ("Errors / 30d", f"{sentry.get('errors_30d', '?')}  (free quota 5,000)"),
            ("Monthly cost", f"${cash:,.2f}" + ("  ← real cash" if cash else " (free)")),
        ]
        y = card("Sentry — Error Monitoring", "🔦", s_rows, y, 32 + len(s_rows) * 15 + 8)

    # footer
    _rect(c, 0, PAGE_H - 26, PAGE_W, 26, fill=PURPLE_DEEP)
    _text(
        c,
        LEFT,
        PAGE_H - 20,
        "AI Identity  ·  Infrastructure Cost Report  ·  Confidential",
        REG,
        8,
        LAVENDER,
    )
    _text(c, LEFT, PAGE_H - 20, label, REG, 8, LAVENDER, center_w=CW)
    c.showPage()
    c.save()
    return path


def build_xlsx(label, summary, flags, neon, atlas, gcp, sentry, history, path):
    import openpyxl
    from openpyxl.styles import Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    purple = "4B2D7F"
    head_fill = PatternFill("solid", fgColor=purple)
    head_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color=purple, size=14)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="E2E0EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, cols):
        for i, h in enumerate(cols, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.fill, cell.font, cell.border = head_fill, head_font, border

    # ── Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"AI Identity — Infra Cost Report · {label}"
    ws["A1"].font = title_font
    ws["A2"] = "Cash out-of-pocket:"
    ws["B2"] = summary.get("out_of_pocket_usd", 0)
    ws["B2"].number_format = "$#,##0.00"
    ws["A3"] = "GCP credit runway:"
    ws["B3"] = summary.get("gcp_credits_total", 0)
    ws["B3"].number_format = "$#,##0.00"
    ws["A4"] = "Atlas run-rate (last full mo.):"
    ws["B4"] = atlas.get("last_invoice_usd", 0)
    ws["B4"].number_format = "$#,##0.00"
    for c in ("A2", "A3", "A4"):
        ws[c].font = bold_font
    hdr(ws, 6, ["Flags / Action Items"])
    for i, f in enumerate(flags, 7):
        ws.cell(row=i, column=1, value=f)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60

    # ── Line Items ──
    ws = wb.create_sheet("Line Items")
    hdr(ws, 1, ["Service", "Item", "Cost (USD)", "Detail"])
    r = 2
    for p in neon.get("projects", []):
        ws.append(["Neon", p["name"], 0.0, f"{p['storage_mb']} MB"])
        r += 1
    for k, v in (atlas.get("service_breakdown") or {}).items():
        ws.append(["Atlas", k, v, atlas.get("last_invoice_period", "")])
        r += 1
    if sentry.get("configured"):
        cash = {"team": 29.0, "business": 80.0}.get((sentry.get("plan") or "").lower(), 0.0)
        ws.append(
            [
                "Sentry",
                f"{(sentry.get('plan') or '?').title()} plan",
                cash,
                f"{sentry.get('errors_30d', '?')} errors/30d",
            ]
        )
        r += 1
    for col, w in zip("ABCD", (12, 36, 14, 22), strict=False):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_col=4):
        row[2].number_format = "$#,##0.00"

    # ── Credits ──
    ws = wb.create_sheet("Credits")
    hdr(ws, 1, ["Pool", "Remaining (USD)", "Expiry", "Note"])
    for cr in gcp.get("credits", []):
        ws.append([cr["name"], cr["usd"], cr["expiry"], cr["note"]])
    ws.append(["Atlas ACCELERATOR-PARTNER-500", "n/a (console-only)", "", "applied at invoicing"])
    for col, w in zip("ABCD", (34, 18, 14, 28), strict=False):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_col=2):
        if isinstance(row[1].value, (int, float)):
            row[1].number_format = "$#,##0.00"

    # ── Trend (month-over-month) ──
    ws = wb.create_sheet("Trend")
    hdr(
        ws,
        1,
        [
            "Month",
            "Cash out-of-pocket",
            "Atlas MTD",
            "Atlas run-rate",
            "GCP credits remaining",
            "Neon storage (MB)",
        ],
    )
    for i, (m, s) in enumerate(sorted(history.items()), 2):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=s.get("out_of_pocket_usd", 0)).number_format = "$#,##0.00"
        ws.cell(row=i, column=3, value=s.get("atlas_mtd", 0)).number_format = "$#,##0.00"
        ws.cell(row=i, column=4, value=s.get("atlas_run_rate", 0)).number_format = "$#,##0.00"
        ws.cell(row=i, column=5, value=s.get("gcp_credits_total", 0)).number_format = "$#,##0.00"
        ws.cell(row=i, column=6, value=s.get("neon_storage_mb", 0))
    for col, w in zip("ABCDEF", (10, 18, 12, 14, 22, 18), strict=False):
        ws.column_dimensions[col].width = w

    wb.save(str(path))
    return path
